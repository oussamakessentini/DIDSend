import shutil
import time
import yaml
import os
from collections.abc import Mapping
import re
import threading
import queue
import xmltodict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
import subprocess
import logging
from typing import Any, List, Optional, Tuple, Union
from pathlib import Path
from Lib.Pdx_Odx import *


# Configure logging
# logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

Global_config_file = "Config.yml"


def format_hex(item):
    if isinstance(item, int):
        return f"{item:02X}"
    elif isinstance(item, (bytes, bytearray)) and len(item) == 1:
        return f"{item[0]:02X}"
    elif len(item) > 1:
        print("Wrong data format passed to format_hex() function")
        return list(item)
    else:
        raise ValueError(f"Cannot format item: {item}")

def is_hex(s):
    try:
        int(s, 16)  # Try converting to an integer with base 16
        return True
    except ValueError:
        return False
    
def is_int(s):
    try:
        int(s)  # Try converting to an integer with base 16
        return True
    except ValueError:
        return False

def isBetween(A, B, C):
    Mi = min(B, C)
    Ma = max(B, C)
    return Mi <= A <= Ma

def dlc_to_data_size(dlc):
    """Convert CAN FD DLC to actual data length."""
    dlc_map = {
        0: 0, 1: 1, 2: 2, 3: 3, 4: 4, 5: 5, 6: 6, 7: 7, 8: 8,
        9: 12, 10: 16, 11: 20, 12: 24, 13: 32, 14: 48, 15: 64
    }
    return dlc_map.get(dlc, -1)

def get_dlc_for_data_length(data_length):
    """Return the smallest DLC value that can accommodate the given data length in CAN FD."""
    dlc_map = [
        (0, 0), (1, 1), (2, 2), (3, 3), (4, 4), (5, 5), (6, 6), (7, 7), (8, 8),
        (9, 12), (10, 16), (11, 20), (12, 24), (13, 32), (14, 48), (15, 64)
    ]

    for dlc, size in dlc_map:
        if data_length <= size:
            return dlc  # Return the DLC that fits the data length

    return -1  # If the length is out of range

def str_to_hexList(strData, symbol=''):
    data = []
    def is_hex_string(s: str) -> bool:
        if len(s) % 2 != 0:
            return False  # Hex data must be even-length for byte pairs
        try:
            bytes.fromhex(s)  # Will raise ValueError if not valid hex
            return True
        except ValueError:
            return False
        
    if len(strData) > 1:
        if(symbol != ''):
            for group in strData.split(symbol):
                if group:
                    data.extend([int(x, 16) for x in group.split() if x])
        else:
            if is_hex_string(strData):
                # Treat as hex string and convert every 2 characters into a byte
                data = [int(strData[i:i+2], 16) for i in range(0, len(strData), 2)]
            else:
                # Treat as regular string and convert each character to its ASCII value
                data = [ord(intData) for intData in strData]
    else:
        data = [int(strData)]
    return data

def load_config(obj_dest, globalVal, config_file, Encode=False):
    """Load configuration from a YAML file."""
    try:
        with open(config_file, "r") as file:
            config = yaml.safe_load(file)
            data = {}
            # Flatten the nested structure and set attributes
            flatten_dict(data, globalVal, config, Encode=Encode)

            for item, value in data.items():
                if isinstance(obj_dest, Mapping) and obj_dest.get(item, "") is None:
                    obj_dest[item] = value
                elif hasattr(obj_dest, item) and getattr(obj_dest, item) is None:
                    setattr(obj_dest, item, value)
    except Exception as e:
        print(f"Cannot Access Config file {config_file} Exception: {str(e)}" )
        config = {}
    return config


def get_nested_yaml_option(
    file_path: Union[str, Path],
    keys: Union[str, List[str]],
    default: Any = None,
    raise_on_missing: bool = False,
    safe_mode: bool = True
) -> Optional[Any]:
    """
    Retrieve a nested option value from a YAML file with support for multiple levels.

    Args:
        file_path: Path to the YAML file (string or Path object)
        keys: Either a single key string (for top-level) or list of keys for nested access
              (e.g., ['category', 'subcategory', 'option'])
        default: Default value to return if option is not found
        raise_on_missing: If True, raises KeyError when any key in the path is missing
        safe_mode: If True, uses yaml.safe_load for security

    Returns:
        The requested value if found, otherwise the default value

    Raises:
        FileNotFoundError: If the YAML file doesn't exist
        PermissionError: If file cannot be read
        yaml.YAMLError: If the YAML file is malformed
        KeyError: If raise_on_missing is True and any key in path is missing
        TypeError: If keys argument is invalid
    """
    try:
        # Validate and normalize input
        if isinstance(keys, str):
            keys = [keys]
        elif not isinstance(keys, list) or not all(isinstance(k, str) for k in keys):
            raise TypeError("keys must be a string or list of strings")

        if not keys:
            raise ValueError("keys list cannot be empty")

        # Convert to Path object immediately
        path = Path(file_path)

        # Verify file exists and is readable
        if not path.exists():
            raise FileNotFoundError(f"YAML file not found: {path}")
        if not path.is_file():
            raise ValueError(f"Path is not a file: {path}")
        if not os.access(path, os.R_OK):
            raise PermissionError(f"Cannot read file: {path}")

        # Load YAML content
        with path.open('r', encoding='utf-8') as file:
            try:
                loader = yaml.safe_load if safe_mode else yaml.load
                data = loader(file) or {}
            except yaml.YAMLError as e:
                logger.error(f"Error parsing YAML file {path}: {e}")
                raise
            except Exception as e:
                logger.error(f"Unexpected error loading YAML: {e}")
                raise

        # Navigate through nested structure
        current_level = data
        for key in keys:
            if not isinstance(current_level, dict):
                if raise_on_missing:
                    raise KeyError(f"Path segment '{key}' not found (reached non-dict level)")
                logger.debug(f"Path segment '{key}' not found (reached non-dict level)")
                return default

            if key not in current_level:
                if raise_on_missing:
                    raise KeyError(f"Key '{key}' not found in path {keys[:keys.index(key)+1]}")
                logger.debug(f"Key '{key}' not found, returning default")
                return default

            current_level = current_level[key]

        return current_level

    except Exception as e:
        logger.error(f"Error retrieving YAML option: {e}")
        if raise_on_missing:
            raise
        return default
    
def find_and_update_key(data, key, new_value):
    """
    Recursively search for a key in a nested dictionary and update its value.
    Returns True if the key was found and updated, otherwise False.
    """
    if isinstance(data, dict):
        for k, v in data.items():
            if k == key:
                data[k] = new_value
                return True  # Stop searching once updated
            elif isinstance(v, dict):  # Recursively search in nested dicts
                if find_and_update_key(v, key, new_value):
                    return True
            elif isinstance(v, list):  # Handle lists of dictionaries
                for item in v:
                    if isinstance(item, dict) and find_and_update_key(item, key, new_value):
                        return True
    return False  # Key not found

def update_config_value(config_file, key, new_value):
    """
    Load the YAML file, update the given key wherever it is found, afnd save the changes.
    """
    try:
        yaml.preserve_quotes = True
        with open(config_file, "r") as file:
            config = yaml.safe_load(file) or {}

        if find_and_update_key(config, key, new_value):
            with open(config_file, "w") as file:
                yaml.dump(config, file, default_flow_style=False, sort_keys=False)
            print(f"Updated '{key}' to {new_value} in {config_file}")
        else:
            print(f"Key '{key}' not found in {config_file}")

    except Exception as e:
        print(f"Error updating YAML: {e}")

def flatten_dict(obj_dest, globalVal, dictionary, Encode=False):
    """Recursively flatten a nested dictionary into class attributes."""
    for key, value in dictionary.items():
        try:
            if isinstance(value, dict):
                flatten_dict(obj_dest, globalVal, value, Encode=Encode)  # Recursive call for nested dicts
            else:
                if isinstance(value, list):
                    val = value
                elif value in globalVal:
                    val = globalVal[value]
                else:
                    val = value.encode() if Encode and isinstance(value, str) else value
                obj_dest[key] = val
        except Exception as e:
            print(f"Cannot set var Exception: {str(e)}" )

def loadConfigFilePath(localPath=""):
    with open(os.path.join(localPath, Global_config_file), "r") as file:
        config = yaml.safe_load(file)
        path_fileConfig = os.path.join(localPath, config["configFile"])
        return path_fileConfig

def verifyFrame(dataReceived, dataSent, size):
    resultStatus = False
    if ((dataReceived[0] & dataSent[0]) == dataSent[0]) and ((dataReceived[0] & 0x40) == 0x40):
        resultStatus = True
        for i in range(1, size):
            if (dataReceived[i] != dataSent[i]):
                resultStatus = False
    return resultStatus

def remove_namespace(tree):
    root = tree.getroot()
    namespace = root.tag.split('}')[0].strip('{') if '}' in root.tag else ''
    for elem in root.iter():
        elem.tag = re.sub(r'\{.*?\}', '', elem.tag)  # Remove namespace
    return tree, namespace

def restore_namespace(tree, namespace):
    if namespace:
        root = tree.getroot()
        for elem in root.iter():
            elem.tag = f'{{{namespace}}}' + elem.tag
    return tree

def find_recursive(element, tag):
    """Recherche récursive du premier élément avec le tag donné."""
    found = element.find(tag)
    if found is not None:
        return found
    for child in element:
        found = find_recursive(child, tag)
        if found is not None:
            return found
    return None

def find_recursive_Value(element, tag, value):
    """Recherche récursive du premier élément avec le tag et valeur donné."""
    found = element.find(tag)
    if found is not None and value in found.text:
        value = element.find("VALUE")
        if value is not None:
            return value.text
        else:
            value = element.find("VALUE-REF")
            if value is not None:
                return value.text
            else:
                return None
    for child in element:
        found = find_recursive_Value(child, tag, value)
        if found is not None:
            return found
    return None

class PeekableQueue(queue.Queue):
    def __init__(self):
        super().__init__()
        self.peek_lock = threading.Lock()  # Lock for safe peeking

    def peek(self):
        """Safely peek at the first item without removing it."""
        with self.peek_lock:  # Prevent other threads from peeking at the same time
            if not self.empty():
                return self.queue[0]  # Access the internal queue directly
            return None  # Return None if the queue is empty

def arxml_to_dict_xmltodict(file_path):
    with open(file_path, "r", encoding="utf-8") as file:
        return xmltodict.parse(file.read())

def wait_ms(ms: int):
    """Wait for the given duration in milliseconds without blocking other threads."""
    time.sleep(ms / 1000.0)

# Checksum CRC-16-CCITT (Polynomial 0x1021)
def crc16_ccitt(data):
    poly = 0x1021 # ✅ Polynomial 0x1021
    crc = 0x0000  # ✅ Initial value 
    for b in data:
        crc ^= b << 8
        for _ in range(8):
            if crc & 0x8000:
                crc = (crc << 1) ^ poly
            else:
                crc <<= 1
            crc &= 0xFFFF
    return crc

def crc16_x25(data: bytes) -> int:
    """
    Calculate CRC-16/X-25 (DECT-R) checksum.
    
    Parameters:
        data: Input data as bytes
        
    Returns:
        16-bit CRC checksum (int)
    """
    poly = 0x1021  # ✅ Polynomial (0x1021, but we use reversed 0x8408)
    crc = 0xFFFF   # ✅ Initial value (will be inverted to 0x0000 after first step)
    
    for byte in data:
        crc ^= byte
        for idx in range(8):
            if crc & 1:
                crc = (crc >> 1) ^ 0x8408  # Reversed polynomial
            else:
                crc >>= 1
    
    crc ^= 0xFFFF  # Final XOR
    return crc & 0xFFFF  # Ensure 16-bit result

# ----------------------------------------    
# Excel functions
# ----------------------------------------
def adjustWidth(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Récupérer la lettre de la colonne (A, B, C, ...)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) if max_length < 100 else 100  # Limit cell size
        ws.column_dimensions[column].width = adjusted_width

def applyPainterFormat(excel_file, column):
    # Set the colors for the painter format
    fill_green  = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
    fill_orange = PatternFill(start_color="DE7B12", end_color="DE7B12", fill_type="solid")  # Orange
    fill_red    = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red

    status_colors = {
        "OK": fill_green,
        "ROUTINE_STARTED": fill_green,
        "ROUTINE_FINISHED_OK": fill_green,
        "ROUTINE_IN_PROGRESS": fill_orange,
        "NOK": fill_red,
    }

    # Load Excel file with openpyxl to add painter format rules
    wb = load_workbook(excel_file)

    for sheet in wb.worksheets:
        # Adjust the width of columns to adapt with the content
        adjustWidth(sheet)

        # Add color rules for the 'Status' column
        max_row = sheet.max_row
        cell_range = f"{column}2:{column}{max_row}"

        for status, fill in status_colors.items():
            sheet.conditional_formatting.add(cell_range, CellIsRule(operator='equal', formula=[f'"{status}"'], fill=fill))

    # Save the Excel file with the rules and format painter
    wb.save(excel_file)

# ----------------------------------------    
# ULP (Motorola S-record) functions
# ----------------------------------------
def run_srec_cat(
    srec_cat_path: str,
    input_files: List[Tuple[str, str]],  # [(filename, format), ...]
    output_file: str,
    output_format: str,
    output_options: Optional[List[str]] = None,
    global_options: Optional[List[str]] = None
) -> bool:
    """
    Runs srec_cat.exe with full support for chaining multiple inputs, formats, and advanced options.

    Args:
        srec_cat_path (str): Path to srec_cat.exe.
        input_files (List[Tuple[str, str]]): List of input files with formats (e.g., [('input.bin', 'Binary')]).
        output_file (str): Path to the output file.
        output_format (str): Output format (e.g., 'Intel', 'Srec', 'Binary').
        output_options (List[str], optional): Additional options for output file (e.g., address range, offset).
        global_options (List[str], optional): Additional global options (e.g., --fill, --output-block-size).

    Returns:
        bool: True if successful, False otherwise.
    """
    if not os.path.isfile(srec_cat_path):
        logging.error(f"srec_cat.exe not found: {srec_cat_path}")
        return False

    for file_path, _ in input_files:
        if not os.path.isfile(file_path):
            logging.error(f"Input file not found: {file_path}")
            return False

    command = [srec_cat_path]

    # Add input files and their formats
    for file_path, fmt in input_files:
        command.extend([file_path, f"-{fmt}"])

    # Add global options if provided
    if global_options:
        command.extend(global_options)

    # Add output file and format
    command.extend(['-o', output_file, f'-{output_format}'])

    # Add additional output options if provided
    if output_options:
        command.extend(output_options)

    # Execute the command
    try:
        result = subprocess.run(
            command,
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True
        )
        logging.info("srec_cat executed successfully.")
        return True
    except subprocess.CalledProcessError as e:
        logging.error("srec_cat execution failed.")
        logging.error(f"Command: {' '.join(command)}")
        logging.error(f"Return Code: {e.returncode}")
        logging.error(f"Stdout: {e.stdout}")
        logging.error(f"Stderr: {e.stderr}")
        return False
    except Exception as ex:
        logging.exception(f"Unexpected error while running srec_cat: {ex}")
        return False

# ----------------------------------------    
# PDX functions
# ----------------------------------------
def get_all_files_path(
    folder_path: str, 
    allowed_extensions: Optional[List[str]] = None,
    exclude_hidden: bool = True
) -> List[str]:
    """
    Recursively retrieves all files path in a directory and its subdirectories.
    
    Args:
        folder_path (str): Path to the directory to scan.
        allowed_extensions (List[str], optional): Only include files with these extensions (e.g., ['.txt', '.csv']). 
            If None, all files are included. Defaults to None.
        exclude_hidden (bool): Whether to exclude hidden files (starting with '.'). Defaults to True.
    
    Returns:
        List[str]: List of absolute file paths.
    
    Raises:
        ValueError: If folder_path does not exist or is not a directory.
        PermissionError: If access to the folder is denied.
    """
    # Set up logging
    logging.basicConfig(level=logging.DEBUG)
    logger = logging.getLogger(__name__)
    
    file_paths = []
    
    try:
        # Validate input path
        if not os.path.exists(folder_path):
            raise ValueError(f"Path does not exist: {folder_path}")
        if not os.path.isdir(folder_path):
            raise ValueError(f"Path is not a directory: {folder_path}")
        
        # Walk through directory
        for root, tmp, files in os.walk(folder_path):
            for file in files:
                # Skip hidden files if enabled
                if exclude_hidden and file.startswith('.'):
                    continue
                
                # Check file extension
                file_path = os.path.join(root, file)
                if allowed_extensions:
                    tmp, ext = os.path.splitext(file)
                    if ext.lower() not in allowed_extensions:
                        continue
                
                file_paths.append(os.path.abspath(file_path))
        
        logger.info(f"Found {len(file_paths)} files in {folder_path}")
        return file_paths
    
    except PermissionError as e:
        logger.error(f"Permission denied while accessing {folder_path}: {e}")
        raise
    except Exception as e:
        logger.error(f"Unexpected error while scanning {folder_path}: {e}")
        raise

def remove_extension(file_path: str) -> str:
    """Removes the file extension while preserving the rest of the path.
    
    Args:
        file_path: Input file path (e.g., '/dir/subdir/file.txt')
    
    Returns:
        Path without extension (e.g., '/dir/subdir/file')
    """
    base, _ = os.path.splitext(file_path)
    return base

def extractPdxFileInfo(pdx_file):
    odxC = Pdx_Odx()
    pdxDict = {}
    odxfDataFile = ''
    pdxfBinFile = ''

    PdxFilePath = remove_extension(pdx_file) + '\\'
    # print(PdxFilePath)

    if os.path.exists(PdxFilePath):
        shutil.rmtree(PdxFilePath)
        # print("\nRemove old PDX file\n")
    
    if pdx_file.lower().endswith(".pdx"):
        odxC.pdx_unzip(pdx_file, PdxFilePath)
        odxfDataFile = odxC.getFilePath(PdxFilePath, None, '.odx-f')
        pdxfBinFile  = odxC.getFilePath(PdxFilePath, None, '.bin')
        # print(odxDataFile)
        try:
            pdxDict = odxC.getPdxData(odxfDataFile)
            # print(pdxDict)
        except:
            print("Error : PDX file data extraction")
    else:
        # Clear old PDX read data :
        print("[Warning] Read PDX => PDX file not found =>", pdx_file)

    return pdxfBinFile, odxfDataFile, pdxDict

def swTypeDesc(data: str) -> str:
    if('BOOT' in data):
        return "Boot Software"
    elif('CODE' in data):
        return "Application Software"
    elif('DATA' in data):
        return "Calibration Software"
    else:
        return "Unkown Software"
    
def int_to_byteList(value:int = 0, byte_length:int = 4):
    # Convert to bytes using only required number of bytes
    int_byte_array = value.to_bytes(byte_length, byteorder='big')

    byteList = list(int_byte_array)

    return byteList