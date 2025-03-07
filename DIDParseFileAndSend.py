from UDS.PCANBasic import *
from UDS.UDS_Frame import UDS_Frame
import pandas as pd
from Utils import *
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

PROJECT = 'PR128'

def adjustWidth(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Récupérer la lettre de la colonne (A, B, C, ...)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)  # Ajouter un peu d'espace
        ws.column_dimensions[column].width = adjusted_width

def Pcan_ReadDID(Pcan, did):
    retVal = Pcan.ReadDID(did)

    if is_hex(retVal[0]) == True:
        result = "OK"
        data = ";".join(format_hex(int(x, 16)) for x in retVal)
        Error = ""
    else:
        result = "NOK"
        data = ""
        Error = str(retVal[1])
    return result, data, Error

def Pcan_WriteDID(Pcan, did, dataraw):
    data = [int(x, 16) for x in dataraw.split(";")]
    retVal = Pcan.WriteDID(did, data)
    if retVal[1] == True:
        status = "OK"
        Error = ""
    else:
        status = "NOK"
        Error = str(retVal[2])
    # Retourne un tuple (status, error)
    return status, Error

def parseAndSend(Pcan):
    # Chemin vers le fichier Excel
    excel_file = 'DID_Status.xlsx'

    # Charger les feuilles "DID Read" et "DID Write" dans des DataFrames
    df_read = pd.read_excel(excel_file, sheet_name='DID Read', dtype=str)
    df_write = pd.read_excel(excel_file, sheet_name='DID Write', dtype=str)

    # Parcourir la feuille "DID Read" ligne par ligne
    for index, ligne in df_read.iterrows():
        # Appeler la fonction Pcan.ReadDID avec le DID de la ligne
        resultat, data, error = Pcan_ReadDID(Pcan, ligne['DID'])
        
        # Mettre à jour la ligne avec les résultats
        df_read.at[index, 'Resultat'] = resultat
        df_read.at[index, 'Data'] = data
        df_read.at[index, 'Error'] = error

    # Parcourir la feuille "DID Write" ligne par ligne
    for index, ligne in df_write.iterrows():
        # Appeler la fonction Pcan.WriteDID avec le DID et la Data de la ligne
        status, error = Pcan_WriteDID(Pcan, ligne['DID'], ligne['Data'])
        
        # Mettre à jour la ligne avec les résultats
        df_write.at[index, 'Status'] = status
        df_write.at[index, 'Error'] = error

    # Sauvegarder les modifications dans le même fichier Excel
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='w') as writer:
        df_read.to_excel(writer, sheet_name='DID Read', index=False)
        df_write.to_excel(writer, sheet_name='DID Write', index=False)

    # Charger le fichier Excel avec openpyxl pour ajouter des règles de mise en forme
    wb = load_workbook(excel_file)
    ws_read = wb['DID Read']
    ws_write = wb['DID Write']

    # Ajuster la largeur des colonnes pour s'adapter au contenu
    adjustWidth(ws_read)
    adjustWidth(ws_write)

    # Définir les couleurs pour la mise en forme conditionnelle
    fill_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Vert
    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # Rouge

    # Ajouter une règle de mise en forme conditionnelle pour la colonne "Resultat"
    # Règle 1 : Si la valeur est "OK", appliquer le vert
    ws_read.conditional_formatting.add(
        f"B2:B{ws_read.max_row}",  # Appliquer à la colonne B (Resultat) à partir de la ligne 2
        CellIsRule(operator='equal', formula=['"OK"'], fill=fill_green)
    )
    ws_write.conditional_formatting.add(
        f"B2:B{ws_write.max_row}",  # Appliquer à la colonne B (Resultat) à partir de la ligne 2
        CellIsRule(operator='equal', formula=['"OK"'], fill=fill_green)
    )

    # Règle 2 : Si la valeur est "NOK", appliquer le rouge
    ws_read.conditional_formatting.add(
        f"B2:B{ws_read.max_row}",  # Appliquer à la colonne B (Resultat) à partir de la ligne 2
        CellIsRule(operator='equal', formula=['"NOK"'], fill=fill_red)
    )
    ws_write.conditional_formatting.add(
        f"B2:B{ws_write.max_row}",  # Appliquer à la colonne B (Resultat) à partir de la ligne 2
        CellIsRule(operator='equal', formula=['"NOK"'], fill=fill_red)
    )

    # Sauvegarder le fichier Excel avec les règles de mise en forme
    wb.save(excel_file)

    print(f"{excel_file} updated successfully")

if __name__ == "__main__":

    if(PROJECT == 'PR105'):
        TxId = 0x6B4
        RxId = 0x694
        Pcan = UDS_Frame(PCAN_USBBUS1, False, PCAN_BAUD_500K, TxId, RxId, False, True)

        # print(Pcan.getFrameFromId(596))
        Pcan.StartSession(3)
        # Pcan.StartReset(2)
        # Pcan.StartReset(3)
        # print(Pcan.ReadDID("F41C"))
        # print(Pcan.ReadDID("D863"))
        # print(Pcan.ReadDID("0101"))
        parseAndSend(Pcan)
        
    else:
        TxId = 0x18DADBF1
        RxId = 0x18DAF1DB
        Pcan = UDS_Frame(PCAN_USBBUS1, False, PCAN_BAUD_500K, TxId, RxId, True, True)

        Pcan.StartSession(3)
        # print(Pcan.ReadDID("8281"))
        parseAndSend(Pcan)
